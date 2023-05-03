import openpyxl

wb = openpyxl.load_workbook('inventory.xlsx')
ws = wb['Sheet1']

supplier_list = {}
total_inventory_value = {}

ws_result = wb.create_sheet('Inventory Value')
ws_result.append(['Product No', 'total inventory value', 'Supplier'])

for row in range(2, ws.max_row + 1):
    product_no = int(ws.cell(row, 1).value)
    inventory = ws.cell(row, 2).value
    price = ws.cell(row, 3).value
    supplier = ws.cell(row, 4).value
    total_value = int(inventory * price)

    # Count the number of suppliers
    if supplier in supplier_list:
        supplier_list[supplier] += 1
    else:
        print('Add a new supplier')
        supplier_list[supplier] = 1

    # List products of inventory less than 10
    if inventory < 10:
        print(f'product {product_no} has inventory less than 10')

    # List each company with respective total inventory value
    if total_inventory_value.get(supplier) is None:
        total_inventory_value[supplier] = {}
    total_inventory_value[supplier][product_no] = total_value

    # Create a new sheet with the following headers: Product No, total inventory value, Supplier
    # Calculate the total inventory value for each product and write it to the new sheet
    ws_result.append([product_no, total_value, supplier])

print(supplier_list)
print(total_inventory_value)

wb.save('inventory_value.xlsx')






